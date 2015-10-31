#!/bin/evn python
#coding: UTF-8
#
#

import codecs
from openpyxl.reader.excel import load_workbook
import os
from types import *
import sys
import codecs
import time


def g_nonone(i,j,table):
    gvalue = table.cell(row = i,column = j).value
    if gvalue==None:
        gvalue=""
    return gvalue

def sheet2list(path):
	wb = load_workbook(path)
	define_name_m = wb.get_sheet_names()[0]
	sheet = wb.get_sheet_by_name(define_name_m)
	max_row = sheet.get_highest_row()
	max_column = sheet.get_highest_column()

	rows = []
	for i in range(2, max_row):
		cells = []
		for j in range(max_column):
			v = g_nonone(i, j, sheet)
			cells.append(v)
		rows.append(cells)
	return rows


